import math
import openpyxl


def generate_workout_plan(one_rep_maxes, num_days):
    # Calculate the weight to use for each lift based on the 1 rep max
    weights = {
        "bench press": one_rep_maxes["bench press"] * 0.65,
        "deadlift": one_rep_maxes["deadlift"] * 0.75,
        "squat": one_rep_maxes["squat"] * 0.75,
        "overhead press": one_rep_maxes["overhead press"] * 0.65,
    }

    # Calculate the number of sets and reps for each lift
    sets_and_reps = {
        "bench press": (4, 6),
        "deadlift": (4, 6),
        "squat": (4, 6),
        "overhead press": (4, 6),
    }

    # Calculate the rest time between sets (in seconds)
    rest_time = {
        "bench press": 60,
        "deadlift": 120,
        "squat": 90,
        "overhead press": 60,
    }

    # Create a list to store the workout plan
    workout_plan = []

    # Generate the workout plan for each day
    for day in range(num_days):
        # Choose the lifts to include in the workout
        lifts = []
        if day % 2 == 0:
            lifts = ["bench press", "deadlift", "squat"]
        else:
            lifts = ["squat", "overhead press", "bench press"]

        # Create a dictionary to store the workout for the day
        workout = {}
        workout["day"] = day + 1
        workout["lifts"] = []

        # Add the lifts to the workout
        for lift in lifts:
            sets, reps = sets_and_reps[lift]
            weight = weights[lift]
            rest = rest_time[lift]
            workout["lifts"].append((lift, sets, reps, weight, rest))

        # Add the workout to the plan
        workout_plan.append(workout)

    return workout_plan


def export_to_excel(workout_plan):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Activate the first sheet
    sheet = wb.active

    # Add the header row
    sheet.cell(row=1, column=1).value = "Day"
    sheet.cell(row=1, column=2).value = "Lift"
    sheet.cell(row=1, column=3).value = "Sets"
    sheet.cell(row=1, column=4).value = "Reps"
    sheet.cell(row=1, column=5).value = "Weight (lbs)"
    sheet.cell(row=1, column=6).value = "Rest (seconds)"

    # Add the workout plan to the sheet
    for row, workout in enumerate(workout_plan, start=2):
        for lift in workout["lifts"]:
            sheet.cell(row=row, column=1).value = workout["day"]
            sheet.cell(row=row, column=2).value = lift[0]
            sheet.cell(row=row, column=3).value = lift[1]
            sheet.cell(row=row, column=4).value = lift[2]
            sheet.cell(row=row, column=5).value = lift[3]
            sheet.cell(row=row, column=6).value = lift[4]
            row += 1

    # Save the workbook
    wb.save("workout_plan.xlsx")
